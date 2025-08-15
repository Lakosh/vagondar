from django.shortcuts import render
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView, DeleteView, View, CreateView
from django.shortcuts import get_object_or_404, redirect
from .models import TrainEvent, TariffHistory, Wagon, Destination
from django.views.generic.edit import FormMixin, UpdateView
from django.db.models import Q
from . import forms
from django.urls import reverse_lazy
from django.http import HttpResponse
from openpyxl import *
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
import io
import math


# class AddWagonView(CreateView):
#     model = Wagon
#     form_class = forms.AddWagonForm
#     template_name = "main/add_wagon_page.html"
#     success_url = reverse_lazy('main')


class MainViews(LoginRequiredMixin, FormMixin, ListView):
    model = TrainEvent
    template_name = "main/index.html"
    context_object_name = 'events'
    ordering = ['-event_date', '-event_time']
    form_class = forms.AddWagonForm
    success_url = reverse_lazy("main")
    paginate_by = 25

    def get_queryset(self):
        query = self.request.GET.get('search_events')
        qs = TrainEvent.objects.select_related("tariff").prefetch_related('wagons').order_by("-event_date", "-event_time")

        if query:
            qs = qs.filter(
                Q(id__icontains=query) |
                Q(wagons__wagon_number__icontains=query) |
                Q(event_date__icontains=query)
            ).distinct()
        return qs

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context.setdefault("add_wagon_form", forms.AddWagonForm)
        context.setdefault("add_event_form", forms.AddEventForm)
        context.setdefault("add_destination_form", forms.AddDestinationForm)
        context.setdefault("add_tariff_form", forms.AddTariffForm)

        for i in context["events"]:
            total_vol_list = []
            total_price = 0

            wagons = i.wagons.all()
            distances = wagons.values_list("destination__distance_km", flat=True)
            min_distance = float(min(distances)) if distances else 0
            for t in wagons:
                dist = t.destination.distance_km
                if dist > min_distance:
                    t.distance_1 = min_distance
                    t.distance_2 = float(dist) - min_distance
                    t.distance_3 = 0
                else:
                    t.distance_1 = dist
                    t.distance_2 = 0
                    t.distance_3 = 0

                t.distance_1 = float(t.distance_1)
                t.distance_2 = float(t.distance_2)
                t.distance_3 = float(t.distance_3)
                t.dist_sum = t.distance_1 + t.distance_2 + t.distance_3
                wagons_count = len(i.wagons.all())
                if wagons_count < 2:
                    t.vol_per_loko =(t.distance_1 * 3) / wagons_count + t.distance_2 * 3 / wagons_count
                else:
                    # vol_per_loko объем за локомотив
                    t.vol_per_loko = (t.distance_1 * 3) / wagons_count + t.distance_2 * 3 / (wagons_count - 1)
                t.total_vol_per_wagon = t.vol_per_loko + t.dist_sum
                temp_price = t.total_vol_per_wagon * float(i.tariff.tariff)
                t.wagon_p_sum = temp_price
                total_price += temp_price
                total_vol_list.append(t.total_vol_per_wagon)
            i.total_vol = sum(total_vol_list)
            i.total_price = total_price

        return context

    def post(self, request, *args, **kwargs):
        form_control = {"add_wagon": forms.AddWagonForm,
                        "add_event": forms.AddEventForm,
                        "add_destination": forms.AddDestinationForm,
                        "add_tariff": forms.AddTariffForm}

        action_type = request.POST.get('action_type')

        if action_type == "edit_event":
            event_id = request.POST.get('event_id')
            event = get_object_or_404(TrainEvent, pk=event_id)
            form = forms.AddEventForm(request.POST, instance=event)
            if form.is_valid():
                form.save()
                return redirect(self.get_success_url())
            else:
                return self.form_invalid(form)

        elif action_type == "edit_wagon":
            wagon_id = request.POST.get("wagon_id")
            wagon = get_object_or_404(Wagon, pk=wagon_id)
            form = forms.AddWagonForm(request.POST, instance=wagon)
            if form.is_valid():
                form.save()
                return redirect(self.get_success_url())
            else:
                return self.form_invalid(form)

        form_class = form_control.get(action_type)
        form = form_class(request.POST)

        if action_type == "add_wagon":
            event_id = request.POST.get('event_id')

            if form.is_valid():
                instance = form.save(commit=False)
                if event_id:
                    instance.event = TrainEvent.objects.get(id=event_id)
                instance.save()
                return redirect(self.get_success_url())
            else:
                return self.form_invalid(form)

        elif action_type in ["add_event", "add_destination", "add_tariff"]:
            if form.is_valid():
                instance = form.save(commit=False)
                instance.save()
                return redirect(self.get_success_url())
            else:
                return self.form_invalid(form)

    def get(self, request, *args, **kwargs):
        if request.GET.get("export") == "1":
            return self.export_to_excel()
        elif request.GET.get("export") == "2":
            return self.export_to_excel_by_config()
        return super().get(request, *args, **kwargs)

    def export_to_excel_by_config(self):
        start_date = self.request.GET.get("start_date")
        end_date = self.request.GET.get("end_date")
        operation_type = self.request.GET.get("operation_type")
        print(start_date, end_date, operation_type, "<=================")
        qs = TrainEvent.objects.select_related("tariff").prefetch_related('wagons').filter(event_date__range=(start_date, end_date)).filter(operation_type=operation_type)

        for i in qs:
            wagons = i.wagons.all()
            print(wagons.values_list("wagon_number", flat=True))

        return HttpResponse("12")

    def export_to_excel(self):
        event_id = self.request.GET.get('event_id')
        qs = TrainEvent.objects.select_related("tariff").prefetch_related('wagons').filter(id=event_id)
        event = qs.first()
        wagons = event.wagons.all()
        wb = Workbook()
        ws = wb.active
        ws.title = "Тестовая страница"

        #  Создание таблицы
        ws.column_dimensions['A'].width = 4
        ws.column_dimensions['B'].width = 39
        ws.column_dimensions['C'].width = 4
        ws.column_dimensions['D'].width = 13
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 21
        ws.column_dimensions["G"].width = 12
        ws.column_dimensions["H"].width = 12
        ws.column_dimensions["I"].width = 12
        ws.column_dimensions["J"].width = 12
        ws.column_dimensions["K"].width = 12
        ws.column_dimensions["L"].width = 12
        ws.column_dimensions["M"].width = 15
        ws.column_dimensions["N"].width = 17
        ws.column_dimensions["O"].width = 13

        ws.merge_cells('B1:C2')
        ws.merge_cells('D1:D2')
        ws.merge_cells('E1:E2')
        ws.merge_cells('F1:F2')
        ws.merge_cells('D1:D2')
        ws.merge_cells('G1:I2')
        ws.merge_cells('J1:J2')
        ws.merge_cells('K1:K2')
        ws.merge_cells('L1:L2')
        ws.merge_cells('M1:M2')
        ws.merge_cells('N1:O2')

        font = Font(name="TimesNewRoman", size=12, bold=True, color="000000")
        font_wagons = Font(name="TimesNewRoman", size=12, bold=False, color="000000")
        alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        thin = Side(border_style="thin", color="000000")

        for border_cords in range(2, 16):
            ws.cell(row=1, column=border_cords).border = Border(top=thin, left=thin, bottom=thin, right=thin)
            ws.cell(row=2, column=border_cords).border = Border(top=thin, left=thin, bottom=thin, right=thin)
            ws.cell(row=1, column=border_cords).font = font
            ws.cell(row=1, column=border_cords).alignment = alignment

        ws["O3"].alignment = alignment

        ws["D1"] = "проезд локомотива (1,5*2)"
        ws["E1"] = "Подача"
        ws["F1"] = "Подача вагона+локомотив"
        ws["G1"] = "Протяженность"
        ws["J1"] = "Обьем за вагон"
        ws["K1"] = "Обьем за локоматив"
        ws["L1"] = "Итого обьем"
        ws["M1"] = 'Тариф ТОО "Е.Ж.Д."'
        ws["N1"] = "Итого"

        all_wagons = event.wagons.all()
        operation_type = event.operation_type
        if operation_type == "arrival":
            operation_type = "Пригон"
        elif operation_type == "removal":
            operation_type = "Уборка"

        event_datetime_info = f"{event.event_date.strftime("%d.%m.%Y")} ({operation_type} {event.event_time.strftime('%H:%M')})"

        distances = list(wagons.values_list('destination__distance_km', flat=True))
        min_distances = float(min(distances)) if distances else distances
        wagons_count = len(all_wagons)

        ws["B1"] = event_datetime_info

        for wagonId in range(wagons_count):
            ready_data = all_wagons[wagonId]
            ws.cell(row=wagonId + 3, column=1).value = wagonId + 1
            ws.cell(row=wagonId + 3, column=2).value = f"{ready_data.destination.name}: {ready_data.wagon_number}"
            ws.cell(row=wagonId + 3, column=2).font = font_wagons
            ws.cell(row=wagonId + 3, column=3).font = font_wagons
            ws.cell(row=wagonId + 3, column=3).value = 1
            ws.cell(row=wagonId + 3, column=4).font = font_wagons
            ws.cell(row=wagonId + 3, column=4).value = 3
            ws.cell(row=wagonId + 3, column=5).font = font_wagons
            ws.cell(row=wagonId + 3, column=5).value = 1
            ws.cell(row=wagonId + 3, column=6).font = font_wagons
            ws.cell(row=wagonId + 3, column=6).value = "=C3+D3*E3"
            ws.cell(row=wagonId + 3, column=7).font = font_wagons
            ws.cell(row=wagonId + 3, column=7).value = min_distances
            ws.cell(row=wagonId + 3, column=8).font = font_wagons
            ws.cell(row=wagonId + 3, column=8).value = float(ready_data.destination.distance_km) - min_distances
            ws.cell(row=wagonId + 3, column=10).font = font_wagons
            ws.cell(row=wagonId + 3, column=10).value = f"=SUM(G{3+wagonId}:I{3+wagonId})"
            ws.cell(row=wagonId + 3, column=10).number_format = "0.000"
            ws.cell(row=wagonId + 3, column=11).font = font_wagons
            ws.cell(row=wagonId + 3, column=11).value = f"=D{3+wagonId}*G{3+wagonId}/{wagons_count}+D{3+wagonId}*H{3+wagonId}/{(wagons_count - 1) if wagons_count > 1 else wagons_count}"
            ws.cell(row=wagonId + 3, column=11).number_format = "0.000"
            ws.cell(row=wagonId + 3, column=12).font = font_wagons
            ws.cell(row=wagonId + 3, column=12).value = f"=SUM(J{3 + wagonId}:K{3 + wagonId})"
            ws.cell(row=wagonId + 3, column=12).number_format = "0.000"
            ws.cell(row=wagonId + 3, column=13).font = font_wagons
            ws.cell(row=wagonId + 3, column=13).value = ready_data.event.tariff.tariff
            ws.cell(row=wagonId + 3, column=14).font = font_wagons
            ws.cell(row=wagonId + 3, column=14).value = f"=L{3+wagonId}*M{3+wagonId}"
            ws.cell(row=wagonId + 3, column=14).number_format = "0.00"
            ws.cell(row=3, column=15).font = font_wagons
            ws.cell(row=3, column=15).value = f"=SUM(N3:N{wagons_count + 2})"
            ws.cell(row=3, column=15).number_format = "0.00"
            ws.merge_cells(f'O3:O{wagons_count+2}')

        for border_x in range(1, wagons_count + 1):
            for border_range in range(2, 16):
                ws.cell(row=2+border_x, column=border_range).border = Border(top=thin, left=thin, bottom=thin, right=thin)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"test{event.id}.xlsx"
        resp = HttpResponse(buffer.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        return resp

        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        resp["Content-Length"] = buffer.getbuffer().nbytes


def edit_event(request, pk):
    event = get_object_or_404(TrainEvent, pk=pk)
    if request.method == 'POST':
        form = AddEventView(request.POST, instance=event)
        if form.is_valid():
            form.save()
            return redirect('main')
    return redirect('main')


class AddEventView(LoginRequiredMixin, ListView):
    model = Wagon
    template_name = "main/add_event_page.html"


class TariffHistoryViews(LoginRequiredMixin, ListView):
    model = TariffHistory
    template_name = "main/tariff_history_page.html"


class DeleteWagonView(View):
    def post(self, request, pk):
        wagon = get_object_or_404(Wagon, pk=pk)
        wagon.delete()
        return redirect('main')


class DeleteEventView(View):
    def post(self, request, pk):
        event = get_object_or_404(TrainEvent, pk=pk)
        event.delete()
        return redirect('main')
